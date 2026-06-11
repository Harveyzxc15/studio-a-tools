#!/usr/bin/env python3
# TOOL: 每日追蹤主機項目（北一＋北二）
"""
每日追蹤主機項目 — 北一區＋北二區 合併版
  輸出單一 Excel，四個分頁：
    北一-每日追蹤主機 / 北一-教育價佔比 / 北二-每日追蹤主機 / 北二-教育價佔比

優化：
  1. 區間 & 月累積 + 兩區查詢全部用 ThreadPoolExecutor 平行跑
  2. 主機台數 mac/iphone/ipad/watch 合併成 1 次查詢（GROUP BY cat4_id）
  3. 教育價 Step2 mac/ipad/watch 合併成 1 次查詢（GROUP BY cat4_id）
  + 執行進度條、總耗時統計

用法：
  python3 warranty_report_all.py            # 自動：最近完整週（週日～週六）
  python3 warranty_report_all.py 5/31 6/4   # 指定起迄日（同年）
  python3 warranty_report_all.py 2026-05-31 2026-06-04
"""
import subprocess, sys, os, glob, io, re, warnings, time, json
import email as _eml
from email.header import decode_header as _decode_hdr
from datetime import date, timedelta, datetime as _dt
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# ── 連線/路徑設定（硬編碼，與 保固搭售率/warranty_report.py 一致）──
JAVA        = "/Library/Java/JavaVirtualMachines/jdk1.8.0_251.jdk/Contents/Home/bin/java"
EPB_CP      = "/Users/bob.1469/Desktop/北一區週報-app:/Library/EPBrowser/EPB/Shell/shell.jar:/Library/EPBrowser/EPB/Shell/lib/*"
EPB_CWD     = "/Users/bob.1469/Desktop/北一區週報-app"
MAIL_BASE   = Path("~/Library/Mail/V10").expanduser()
OUTPUT_BASE = Path("~/工具中心/輸出/每日追蹤主機").expanduser()

# 同時併發的 EPB 查詢數（過高可能讓 EPB 伺服器吃力，4 為穩妥值）
MAX_WORKERS = 4

# 信件掃描快取：已定案（SETTLE_DAYS 天前）的 setup 日期存本機，warm run 只重掃最近幾天
SETTLE_DAYS  = 2
CACHE_PATH   = OUTPUT_BASE / 'mysetup_cache.json'

# ── 區域設定 ──────────────────────────────────────────────────────────────────
REGIONS = {
    '北一區': {
        'stores':       ['士林', '微風', '美麗華', '阿波羅', '高島屋', '羅東'],
        'shop_codes':   {'士林':'004','微風':'005','美麗華':'024','阿波羅':'046','高島屋':'054','羅東':'057'},
        'mysetup_part': ['士林', '微風', '美麗華', '阿波羅', '高島屋'],
        'mysetup_kw':   {'士林':'士林','微風':'微風','美麗華':'美麗華','阿波羅':'阿波羅','高島屋':'高島屋'},
        'edu_excl':     {'羅東'},
    },
    '北二區': {
        'stores':       ['永和', '板橋誠品', '西門', '花蓮', '板橋遠百', '新莊宏匯', '新店裕隆城'],
        'shop_codes':   {'永和':'009','板橋誠品':'025','西門':'050','花蓮':'055','板橋遠百':'063','新莊宏匯':'064','新店裕隆城':'068'},
        'mysetup_part': ['西門', '板橋遠百', '新莊宏匯', '新店裕隆城'],
        'mysetup_kw':   {'西門':'西門','板橋遠百':'板橋遠百','新莊宏匯':'宏匯','新店裕隆城':'裕隆城'},
        'edu_excl':     {'花蓮', '板橋誠品', '永和'},
    },
}

# (key, 顯示名, header色, sub色, 目標%)
CATS = [
    ('mac',     'Mac',     '2E75B6', 'BDD7EE', 60),
    ('iphone',  'iPhone',  'C55A11', 'FCE4D6', 30),
    ('ipad',    'iPad',    '375623', 'E2EFDA', 50),
    ('watch',   'Watch',   'A50021', 'FFCCCC', 40),
    ('airpods', 'AirPods', '7030A0', 'E2CFEE', 35),
]
EDU_CATS = [
    ('mac',   'Mac',   '2E75B6', 'BDD7EE'),
    ('ipad',  'iPad',  '375623', 'E2EFDA'),
    ('watch', 'Watch', 'A50021', 'FFCCCC'),
]

CAT4_HOST = {'4001':'mac','4002':'mac','4004':'iphone',
             '4005':'ipad','4006':'ipad','4041':'ipad','4038':'watch'}
CAT4_EDU  = {'4001':'mac','4002':'mac',
             '4005':'ipad','4006':'ipad','4041':'ipad','4038':'watch'}

def shop_in(cfg):
    return ','.join(f"'{c}'" for c in cfg['shop_codes'].values())

# ── 日期解析 ──────────────────────────────────────────────────────────────────
def parse_date(s):
    s = s.strip()
    if len(s) == 10 and s[4] == '-':
        return date.fromisoformat(s)
    if '/' in s:
        m, d = s.split('/')
        return date(date.today().year, int(m), int(d))
    raise ValueError(f"無法解析日期：{s}")

def last_saturday():
    d = date.today()
    offset = (d.weekday() + 2) % 7
    return d - timedelta(days=offset if offset else 7)

def fmt(d):
    return d.strftime('%-m/%-d')

# ── EPB 查詢 ──────────────────────────────────────────────────────────────────
def epb(sql, max_rows=10000):
    r = subprocess.run(
        [JAVA, '-Dsun.net.client.defaultReadTimeout=120000',
         '-cp', EPB_CP, 'EPBReportQuery', sql, str(max_rows)],
        capture_output=True, text=True, cwd=EPB_CWD
    )
    lines = [l for l in r.stdout.strip().split('\n') if l.strip()]
    if not lines:
        return []
    hdrs = [h.strip().upper() for h in lines[0].split('\t')]
    return [dict(zip(hdrs, row.split('\t'))) for row in lines[1:]]

QTY = ("SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
       "WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units")

def query_period(cfg, d_start, d_end):
    si   = shop_in(cfg)
    ds   = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de   = f"TO_DATE('{d_end}','yyyy-mm-dd')"
    cond = f"l.doc_date>={ds} AND l.doc_date<={de} AND l.trans_type IN ('A','E','H')"

    # 主機台數：mac/iphone/ipad/watch 合併一次查詢（排除認證機 cat2=2029）
    host = {c: defaultdict(int) for c in ('mac', 'iphone', 'ipad', 'watch')}
    for r in epb(f"SELECT l.shop_id, l.cat4_id, {QTY} FROM poslinev_bi l "
                 f"WHERE l.org_id='01' AND l.shop_id IN ({si}) AND {cond} "
                 f"AND l.cat2_id<>'2029' "
                 f"AND l.cat4_id IN ('4001','4002','4004','4005','4006','4041','4038') "
                 f"GROUP BY l.shop_id, l.cat4_id"):
        cat = CAT4_HOST.get(r.get('CAT4_ID', '').strip())
        if not cat:
            continue
        try:
            sid = str(int(r['SHOP_ID']))
        except (ValueError, KeyError):
            continue
        host[cat][sid] += int(float(r.get('UNITS', 0) or 0))

    # AirPods 主機（cat6 + cat3=3002，排除認證機）
    air = defaultdict(int)
    for r in epb(f"SELECT l.shop_id, {QTY} FROM poslinev_bi l "
                 f"WHERE l.org_id='01' AND l.shop_id IN ({si}) AND {cond} "
                 f"AND l.cat2_id<>'2029' AND l.cat6_id IN ('6258','6312','6330') AND l.cat3_id='3002' "
                 f"GROUP BY l.shop_id"):
        try:
            air[str(int(r['SHOP_ID']))] += int(float(r.get('UNITS', 0) or 0))
        except (ValueError, KeyError):
            continue
    host['airpods'] = air

    # SACare 台數（cat6 6533-6537）
    C6MAP = {'6533':'mac','6534':'ipad','6535':'iphone','6536':'watch','6537':'airpods'}
    sa = defaultdict(lambda: defaultdict(int))
    for row in epb(f"SELECT l.shop_id, l.cat6_id, {QTY} FROM poslinev_bi l "
                   f"WHERE l.org_id='01' AND l.shop_id IN ({si}) AND {cond} "
                   f"AND l.cat6_id IN ('6533','6534','6535','6536','6537') "
                   f"GROUP BY l.shop_id, l.cat6_id"):
        cat = C6MAP.get(row.get('CAT6_ID', '').strip())
        if not cat:
            continue
        try:
            sa[str(int(row['SHOP_ID']))][cat] += int(float(row.get('UNITS', 0) or 0))
        except (ValueError, KeyError):
            continue

    # ACPP+ 台數（cat3=3032，依名稱分類）
    def acpp_cat(name):
        n = name.lower()
        if any(k in n for k in ['macbook','mac mini','imac','mac studio']): return 'mac'
        if 'iphone' in n: return 'iphone'
        if 'ipad'   in n: return 'ipad'
        if 'watch'  in n: return 'watch'
        if 'airpods'in n: return 'airpods'
        return None
    acpp = defaultdict(lambda: defaultdict(int))
    for row in epb(f"SELECT l.shop_id, l.name, {QTY} FROM poslinev_bi l "
                   f"WHERE l.org_id='01' AND l.shop_id IN ({si}) AND {cond} "
                   f"AND l.cat3_id='3032' GROUP BY l.shop_id, l.name"):
        cat = acpp_cat(row.get('NAME', ''))
        if not cat:
            continue
        try:
            acpp[str(int(row['SHOP_ID']))][cat] += int(float(row.get('UNITS', 0) or 0))
        except (ValueError, KeyError):
            continue

    out = {}
    for s in cfg['stores']:
        c = str(int(cfg['shop_codes'][s]))
        out[s] = {cat: (host[cat].get(c, 0), sa[c][cat], acpp[c][cat]) for cat, *_ in CATS}
    return out

def query_extras(cfg, d_start, d_end):
    si   = shop_in(cfg)
    ds   = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de   = f"TO_DATE('{d_end}','yyyy-mm-dd')"
    cond = f"l.doc_date>={ds} AND l.doc_date<={de} AND l.trans_type IN ('A','E','H')"
    arp_rows = epb(f"SELECT l.shop_id, {QTY} FROM poslinev_bi l "
                   f"WHERE l.org_id='01' AND l.shop_id IN ({si}) "
                   f"AND {cond} AND l.brand_id='496' GROUP BY l.shop_id")
    spk_rows = epb(f"SELECT l.shop_id, SUM(CASE WHEN l.trans_type IN ('A','H') "
                   f"THEN l.line_total_net + l.line_tax "
                   f"WHEN l.trans_type='E' THEN l.line_total_net + l.line_tax ELSE 0 END) AS amt "
                   f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({si}) "
                   f"AND {cond} AND l.cat4_id='4013' AND l.brand_id<>'453' GROUP BY l.shop_id")
    arp = {str(int(r['SHOP_ID'])): int(float(r.get('UNITS', 0) or 0)) for r in arp_rows}
    spk = {str(int(r['SHOP_ID'])): int(round(float(r.get('AMT', 0) or 0))) for r in spk_rows}
    return {s: {'arpedia': arp.get(str(int(cfg['shop_codes'][s])), 0),
                'speaker': spk.get(str(int(cfg['shop_codes'][s])), 0)}
            for s in cfg['stores']}

def query_edu_units(cfg, d_start, d_end):
    """教育價主機台數：同一單據含 SKU 99200202/99200203，且排除認證機(cat2≠2029)"""
    si = shop_in(cfg)
    ds = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de = f"TO_DATE('{d_end}','yyyy-mm-dd')"

    # Step 1: 含教育價 SKU 的 (shop_id, doc_id) 清單
    edu_pairs = set()
    for r in epb(f"SELECT DISTINCT l.shop_id, l.doc_id FROM poslinev_bi l "
                 f"WHERE l.org_id='01' AND l.shop_id IN ({si}) "
                 f"AND l.doc_date>={ds} AND l.doc_date<={de} "
                 f"AND l.stk_id IN ('99200202','99200203')"):
        try:
            edu_pairs.add((str(int(r['SHOP_ID'])), r['DOC_ID'].strip()))
        except (ValueError, KeyError):
            pass
    if not edu_pairs:
        return {'mac': {}, 'ipad': {}, 'watch': {}}

    # Step 2: mac/ipad/watch 主機台數合併一次查詢，Python 端比對 edu_pairs
    result = {'mac': defaultdict(int), 'ipad': defaultdict(int), 'watch': defaultdict(int)}
    for r in epb(f"SELECT l.shop_id, l.doc_id, l.cat4_id, {QTY} FROM poslinev_bi l "
                 f"WHERE l.org_id='01' AND l.shop_id IN ({si}) "
                 f"AND l.doc_date>={ds} AND l.doc_date<={de} AND l.trans_type IN ('A','E','H') "
                 f"AND l.cat2_id<>'2029' AND l.cat4_id IN ('4001','4002','4005','4006','4041','4038') "
                 f"GROUP BY l.shop_id, l.doc_id, l.cat4_id"):
        cat = CAT4_EDU.get(r.get('CAT4_ID', '').strip())
        if not cat:
            continue
        try:
            sid = str(int(r['SHOP_ID']))
        except (ValueError, KeyError):
            continue
        if (sid, r['DOC_ID'].strip()) in edu_pairs:
            result[cat][sid] += int(float(r.get('UNITS', 0) or 0))
    return {k: dict(v) for k, v in result.items()}

# ── Mysetup 信件掃描 ──────────────────────────────────────────────────────────
def _decode_mail_str(s):
    if not s: return ''
    out = []
    for part, enc in _decode_hdr(s):
        out.append(part.decode(enc or 'utf-8', errors='replace') if isinstance(part, bytes) else part)
    return ''.join(out)

def _parse_emlx(raw):
    nl = raw.index(b'\n')
    try:
        n = int(raw[:nl].strip())
        return _eml.message_from_bytes(raw[nl + 1: nl + 1 + n])
    except Exception:
        return _eml.message_from_bytes(raw.split(b'\n', 1)[1])

def _subject_date(subject):
    m = re.search(r'(\d{1,2})-(\d{1,2})月-(\d{2})', subject)
    if not m: return None
    try:
        return date(2000 + int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None

def _load_cache():
    try:
        with open(CACHE_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _save_cache(cache):
    try:
        OUTPUT_BASE.mkdir(parents=True, exist_ok=True)
        with open(CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception:
        pass

def scan_mysetup(kw_map, lo_date, hi_date):
    """掃 Apple Mail 抓 Personal Setup 提交數，一次涵蓋 lo~hi 的所有 setup 日期。
    回傳 {date_iso: {store: count}}，由 mysetup_sum() 推導各期間（避免重複掃描）。
    純 Python CPU 工作（會佔 GIL），主程式以獨立 process 執行使其與 EPB 查詢併行。
    已定案日期（SETTLE_DAYS 天前）的提交數從本機快取讀取，省下 xlsx 載入。
    """
    cache = _load_cache()
    settle_before = date.today() - timedelta(days=SETTLE_DAYS)

    emlx_files = glob.glob(os.path.join(str(MAIL_BASE), '**/Messages/*.emlx'), recursive=True)
    emlx_files = [f for f in emlx_files if not f.endswith('.partial.emlx')]
    lo = _dt(lo_date.year, lo_date.month, lo_date.day).timestamp() - 86400
    hi = _dt(hi_date.year, hi_date.month, hi_date.day).timestamp() + 86400 * 2
    emlx_files = [f for f in emlx_files if lo <= os.path.getmtime(f) <= hi]

    by_date = {}
    seen_dates = set()
    for path in emlx_files:
        try:
            # 快速預篩：先讀檔頭，原始 bytes 不含關鍵字就跳過（省下對數千封非候選信的 MIME parse）
            with open(path, 'rb') as f:
                raw = f.read(32768)
                if b'Personal Setup' not in raw or b'Setup Data' not in raw:
                    continue
                raw += f.read()  # 候選信才補讀完整內容（attachment 需要）
            if b'\n' not in raw: continue
            msg = _parse_emlx(raw)
            subj = _decode_mail_str(msg.get('Subject', ''))
            if 'Personal Setup' not in subj or 'Setup Data' not in subj: continue
            data_date = _subject_date(subj)
            if data_date is None or not (lo_date <= data_date <= hi_date): continue
            if data_date in seen_dates: continue
            seen_dates.add(data_date)
            diso = data_date.isoformat()
            # 已定案 + 已快取 → 直接用，跳過昂貴的 xlsx 載入
            if data_date < settle_before and diso in cache:
                by_date[diso] = cache[diso]
                continue
            counts = defaultdict(int)
            for part in msg.walk():
                fn = _decode_mail_str(part.get_filename() or '')
                if not fn.lower().endswith('.xlsx'): continue
                raw_xls = part.get_payload(decode=True)
                if not raw_xls: continue
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    mwb = load_workbook(io.BytesIO(raw_xls), read_only=True, data_only=True)
                for sheet in mwb.worksheets:
                    headers = None
                    pos_col = None
                    for row in sheet.iter_rows(values_only=True):
                        if headers is None:
                            headers = [str(v).strip() if v else '' for v in row]
                            if 'POS Name' in headers:
                                pos_col = headers.index('POS Name')
                            else:
                                break
                            continue
                        if pos_col is None: break
                        pos = str(row[pos_col]) if row[pos_col] else ''
                        for store, kw in kw_map.items():
                            if kw in pos:
                                counts[store] += 1
                                break
                mwb.close()
                break
            by_date[diso] = dict(counts)
        except Exception:
            continue

    # 回寫快取：只存已定案日期（最近未定案的日期每次重掃，避免漏後到的資料）
    for diso, counts in by_date.items():
        if date.fromisoformat(diso) < settle_before:
            cache[diso] = counts
    _save_cache(cache)
    return by_date

def mysetup_sum(by_date, start, end):
    """從 scan_mysetup 的 {date_iso:{store:count}} 加總指定期間內各門市提交數。"""
    out = defaultdict(int)
    for diso, counts in by_date.items():
        d = date.fromisoformat(diso)
        if start <= d <= end:
            for store, c in counts.items():
                out[store] += c
    return dict(out)

# ── Excel 輸出 ────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BFBFBF')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, r, c, val=None, bg=None, fg='000000', bold=False, sz=10):
    cell = ws.cell(row=r, column=c)
    cell.value = val
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORD
    return cell

def write_table(ws, start_row, title, data, cfg, extras=None, mysetup=None):
    stores      = cfg['stores']
    ms_part     = cfg['mysetup_part']
    mysetup_col = 2 + len(CATS) * 5 + (2 if extras else 0)
    total_cols  = 1 + len(CATS) * 5 + (2 if extras else 0) + (2 if mysetup is not None else 0)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    sc(ws, start_row, 1, title, bg='1F3864', fg='FFFFFF', bold=True, sz=12)
    ws.row_dimensions[start_row].height = 40

    r2 = start_row + 1
    sc(ws, r2, 1, '門市', bg='404040', fg='FFFFFF', bold=True)
    ws.row_dimensions[r2].height = 22
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 4)
        sc(ws, r2, col, f'{label}  (目標 {tgt}%)', bg=hbg, fg='FFFFFF', bold=True, sz=10)
        col += 5
    if extras:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 1)
        sc(ws, r2, col, 'ARpedia / 喇叭', bg='404040', fg='FFFFFF', bold=True, sz=10)
    if mysetup is not None:
        ws.merge_cells(start_row=r2, start_column=mysetup_col, end_row=r2, end_column=mysetup_col + 1)
        sc(ws, r2, mysetup_col, 'Mysetup 提交率', bg='1F4E79', fg='FFFFFF', bold=True, sz=10)

    r3 = start_row + 2
    sc(ws, r3, 1, '', bg='595959')
    ws.row_dimensions[r3].height = 32
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        for h in ['台數', 'ACPP+', 'SACare', '搭售率', '缺口']:
            sc(ws, r3, col, h, bg=sbg, bold=True, sz=9)
            col += 1
    if extras:
        sc(ws, r3, col,     'ARpedia\n數量', bg='D9D9D9', bold=True, sz=9)
        sc(ws, r3, col + 1, '喇叭\n金額',    bg='D9D9D9', bold=True, sz=9)
    if mysetup is not None:
        sc(ws, r3, mysetup_col,     'Mysetup\n提交數', bg='DEEAF1', bold=True, sz=9)
        sc(ws, r3, mysetup_col + 1, '提交率',          bg='DEEAF1', bold=True, sz=9)

    for i, store in enumerate(stores):
        r  = start_row + 3 + i
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F5F5'
        ws.row_dimensions[r].height = 20
        sc(ws, r, 1, store, bg=bg, bold=True)
        col = 2
        for cat, label, hbg, sbg, tgt in CATS:
            host, sa_, acpp_ = data[store][cat]
            w   = sa_ + acpp_
            pct = w / host * 100 if host > 0 else None
            ps  = f'{pct:.0f}%' if pct is not None else '—'
            txt = ('C00000' if pct is not None and pct < tgt else '375623' if pct is not None else 'AAAAAA')
            # 缺口 = 達標所需搭售數(主機×目標%，無條件進位) − 現有搭售數
            need = (host * tgt + 99) // 100 if host > 0 else 0
            gap  = max(0, need - w)
            if host <= 0:
                gtxt, gfg = '—', 'AAAAAA'
            elif gap == 0:
                gtxt, gfg = '達標', '375623'
            else:
                gtxt, gfg = gap, 'C00000'
            sc(ws, r, col,     host if host > 0 else 0,     bg=bg)
            sc(ws, r, col + 1, acpp_ if acpp_ != 0 else '', bg=bg)
            sc(ws, r, col + 2, sa_   if sa_   != 0 else '', bg=bg)
            sc(ws, r, col + 3, ps, bg=bg, fg=txt, bold=True)
            sc(ws, r, col + 4, gtxt, bg=bg, fg=gfg, bold=True)
            col += 5
        if extras:
            sc(ws, r, col,     extras[store]['arpedia'] or '', bg=bg)
            sc(ws, r, col + 1, extras[store]['speaker']  or '', bg=bg)
        if mysetup is not None:
            if store in ms_part:
                ms    = mysetup.get(store, 0)
                denom = sum(data[store][c][0] for c, *_ in CATS if c != 'airpods')
                pv    = ms / denom * 100 if denom > 0 else None
                ps    = f'{pv:.0f}%' if pv is not None else '—'
                txt   = ('C00000' if pv < 60 else '375623') if pv is not None else 'AAAAAA'
                sc(ws, r, mysetup_col,     ms, bg=bg)
                sc(ws, r, mysetup_col + 1, ps, bg=bg, fg=txt, bold=True)
            else:
                sc(ws, r, mysetup_col,     '—', bg=bg, fg='AAAAAA')
                sc(ws, r, mysetup_col + 1, '—', bg=bg, fg='AAAAAA')

    r_tot = start_row + 3 + len(stores)
    ws.row_dimensions[r_tot].height = 22
    sc(ws, r_tot, 1, '合計', bg='404040', fg='FFFFFF', bold=True)
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        t_h = sum(data[s][cat][0] for s in stores)
        t_w = sum(data[s][cat][1] + data[s][cat][2] for s in stores)
        pct = t_w / t_h * 100 if t_h > 0 else None
        ps  = f'{pct:.0f}%' if pct is not None else '—'
        txt = ('C00000' if pct is not None and pct < tgt else '375623' if pct is not None else 'AAAAAA')
        t_need = (t_h * tgt + 99) // 100 if t_h > 0 else 0
        t_gap  = max(0, t_need - t_w)
        gtxt, gfg = ('—', 'AAAAAA') if t_h <= 0 else (('達標', '375623') if t_gap == 0 else (t_gap, 'C00000'))
        sc(ws, r_tot, col,     t_h, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, sum(data[s][cat][2] for s in stores) or '', bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 2, sum(data[s][cat][1] for s in stores), bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 3, ps, fg=txt, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 4, gtxt, fg=gfg, bg='D9D9D9', bold=True)
        col += 5
    if extras:
        sc(ws, r_tot, col,     sum(extras[s]['arpedia'] for s in stores), bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, sum(extras[s]['speaker']  for s in stores), bg='D9D9D9', bold=True)
    if mysetup is not None:
        t_ms    = sum(mysetup.get(s, 0) for s in ms_part)
        t_denom = sum(data[s][c][0] for s in ms_part for c, *_ in CATS if c != 'airpods')
        pv      = t_ms / t_denom * 100 if t_denom > 0 else None
        ps      = f'{pv:.0f}%' if pv is not None else '—'
        txt     = ('C00000' if pv < 60 else '375623') if pv is not None else 'AAAAAA'
        sc(ws, r_tot, mysetup_col,     t_ms, bg='D9D9D9', bold=True)
        sc(ws, r_tot, mysetup_col + 1, ps,   bg='D9D9D9', bold=True, fg=txt)

    return r_tot + 2

def write_edu_table(ws, start_row, title, edu_data, host_data, cfg):
    stores     = cfg['stores']
    shop_codes = cfg['shop_codes']
    edu_excl   = cfg['edu_excl']
    total_cols = 1 + len(EDU_CATS) * 3

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    sc(ws, start_row, 1, title, bg='1F3864', fg='FFFFFF', bold=True, sz=12)
    ws.row_dimensions[start_row].height = 28

    r2 = start_row + 1
    sc(ws, r2, 1, '門市', bg='404040', fg='FFFFFF', bold=True)
    ws.row_dimensions[r2].height = 22
    col = 2
    for cat, label, hbg, sbg in EDU_CATS:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 2)
        sc(ws, r2, col, label, bg=hbg, fg='FFFFFF', bold=True)
        col += 3

    r3 = start_row + 2
    sc(ws, r3, 1, '', bg='595959')
    ws.row_dimensions[r3].height = 28
    col = 2
    for cat, label, hbg, sbg in EDU_CATS:
        sc(ws, r3, col,     '教育台數', bg=sbg, bold=True, sz=9)
        sc(ws, r3, col + 1, '總台數',   bg=sbg, bold=True, sz=9)
        sc(ws, r3, col + 2, '教育佔比', bg=sbg, bold=True, sz=9)
        col += 3

    for i, store in enumerate(stores):
        r  = start_row + 3 + i
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F5F5'
        ws.row_dimensions[r].height = 20
        sc(ws, r, 1, store, bg=bg, bold=True)
        col = 2
        if store in edu_excl:
            for _ in EDU_CATS:
                sc(ws, r, col,     '—', bg=bg, fg='AAAAAA')
                sc(ws, r, col + 1, '—', bg=bg, fg='AAAAAA')
                sc(ws, r, col + 2, '—', bg=bg, fg='AAAAAA')
                col += 3
        else:
            sid = str(int(shop_codes[store]))
            for cat, *_ in EDU_CATS:
                edu_u = edu_data[cat].get(sid, 0)
                tot_u = host_data[store][cat][0]
                pct   = edu_u / tot_u * 100 if tot_u > 0 else None
                ps    = f'{pct:.0f}%' if pct is not None else '—'
                sc(ws, r, col,     edu_u if edu_u > 0 else 0, bg=bg)
                sc(ws, r, col + 1, tot_u if tot_u > 0 else 0, bg=bg)
                sc(ws, r, col + 2, ps, bg=bg, bold=True)
                col += 3

    incl = [s for s in stores if s not in edu_excl]
    r_tot = start_row + 3 + len(stores)
    ws.row_dimensions[r_tot].height = 22
    sc(ws, r_tot, 1, '合計', bg='404040', fg='FFFFFF', bold=True)
    col = 2
    for cat, *_ in EDU_CATS:
        t_edu = sum(edu_data[cat].get(str(int(shop_codes[s])), 0) for s in incl)
        t_tot = sum(host_data[s][cat][0] for s in incl)
        pct   = t_edu / t_tot * 100 if t_tot > 0 else None
        ps    = f'{pct:.0f}%' if pct is not None else '—'
        sc(ws, r_tot, col,     t_edu, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, t_tot, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 2, ps,    bg='D9D9D9', bold=True)
        col += 3

    return r_tot + 2

def set_col_widths(ws):
    ws.column_dimensions['A'].width = 7
    col = 2
    for _ in CATS:
        for w in [6, 7, 7, 7, 6]:
            ws.column_dimensions[get_column_letter(col)].width = w
            col += 1
    ws.column_dimensions[get_column_letter(col)].width     = 9
    ws.column_dimensions[get_column_letter(col + 1)].width = 11
    ws.column_dimensions[get_column_letter(col + 2)].width = 8
    ws.column_dimensions[get_column_letter(col + 3)].width = 7

# ── 進度條 ────────────────────────────────────────────────────────────────────
def render_progress(done, total, label=''):
    filled = int(30 * done / total) if total else 0
    bar = '█' * filled + '░' * (30 - filled)
    pct = int(100 * done / total) if total else 0
    sys.stdout.write(f'\r  [{bar}] {pct:3d}%  ({done}/{total})  {label:<30}')
    sys.stdout.flush()

# ── 主程式 ────────────────────────────────────────────────────────────────────
def main():
    t0 = time.time()
    if len(sys.argv) >= 3:
        d_start, d_end = parse_date(sys.argv[1]), parse_date(sys.argv[2])
    elif len(sys.argv) == 2:
        d_end   = parse_date(sys.argv[1])
        d_start = d_end.replace(day=1)
    else:
        sat     = last_saturday()
        d_start = sat - timedelta(days=6)
        d_end   = sat
    m_start = d_end.replace(day=1)

    print(f"區間：{fmt(d_start)} ～ {fmt(d_end)}　月累積：{fmt(m_start)} ～ {fmt(d_end)}")

    periods = {'range': (d_start, d_end), 'month': (m_start, d_end)}
    all_kw  = {**REGIONS['北一區']['mysetup_kw'], **REGIONS['北二區']['mysetup_kw']}

    # ── EPB 平行任務（mysetup 另開 process）──
    tasks = []  # (key, callable)
    for rname, cfg in REGIONS.items():
        for pname, (ps, pe) in periods.items():
            tasks.append((('period', rname, pname), lambda cfg=cfg, ps=ps, pe=pe: query_period(cfg, ps, pe)))
            tasks.append((('extras', rname, pname), lambda cfg=cfg, ps=ps, pe=pe: query_extras(cfg, ps, pe)))
            tasks.append((('edu',    rname, pname), lambda cfg=cfg, ps=ps, pe=pe: query_edu_units(cfg, ps, pe)))

    results = {}
    total = len(tasks) + 1  # +1 = 信件掃描
    done = 0
    render_progress(0, total, '開始查詢…')

    # 信件掃描丟到獨立 process，與 EPB 池真正併行（避免 GIL 互鎖）；只掃一次涵蓋兩期間
    lo = min(m_start, d_start)
    with ProcessPoolExecutor(max_workers=1) as pex:
        ms_fut = pex.submit(scan_mysetup, all_kw, lo, d_end)
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futs = {ex.submit(fn): key for key, fn in tasks}
            for fut in as_completed(futs):
                key = futs[fut]
                results[key] = fut.result()
                done += 1
                render_progress(done, total, f"{key[0]} {key[1]} {key[2]}")
        render_progress(done, total, '等待信件掃描…')
        by_date = ms_fut.result()
    done += 1
    render_progress(done, total, '信件掃描完成')
    print()

    ms_range = mysetup_sum(by_date, d_start, d_end)
    ms_month = mysetup_sum(by_date, m_start, d_end)

    # ── 組裝 Excel（4 分頁）──
    wb = Workbook()
    wb.remove(wb.active)
    NOTE = '※ 主機台數已排除認證機（類別2=2029）'
    for rname, cfg in REGIONS.items():
        short = rname[:2]
        dr  = results[('period', rname, 'range')]
        dm  = results[('period', rname, 'month')]
        er  = results[('extras', rname, 'range')]
        em  = results[('extras', rname, 'month')]
        edr = results[('edu',    rname, 'range')]
        edm = results[('edu',    rname, 'month')]
        msr = {s: ms_range.get(s, 0) for s in cfg['stores']}
        msm = {s: ms_month.get(s, 0) for s in cfg['stores']}

        ws = wb.create_sheet(f'{short}-每日追蹤主機')
        ws.freeze_panes = 'B4'
        rt = f"區間  {fmt(d_start)} ～ {fmt(d_end)}  各門市保固搭售率（{rname}）\n{NOTE}"
        mt = f"月累積  {fmt(m_start)} ～ {fmt(d_end)}  各門市保固搭售率（{rname}）\n{NOTE}"
        nr = write_table(ws, 1, rt, dr, cfg, er, msr)
        write_table(ws, nr, mt, dm, cfg, em, msm)
        set_col_widths(ws)

        ws2 = wb.create_sheet(f'{short}-教育價佔比')
        ws2.freeze_panes = 'B4'
        rt2 = f"區間  {fmt(d_start)} ～ {fmt(d_end)}  教育價佔比（{rname}）"
        mt2 = f"月累積  {fmt(m_start)} ～ {fmt(d_end)}  教育價佔比（{rname}）"
        nr2 = write_edu_table(ws2, 1, rt2, edr, dr, cfg)
        write_edu_table(ws2, nr2, mt2, edm, dm, cfg)
        ws2.column_dimensions['A'].width = 9
        for i in range(len(EDU_CATS) * 3):
            ws2.column_dimensions[get_column_letter(2 + i)].width = 9

    out_dir = OUTPUT_BASE
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"每日追蹤主機_北一北二_{d_start.strftime('%m%d')}-{d_end.strftime('%m%d')}.xlsx"
    out_path = out_dir / fname
    wb.save(out_path)
    print(f"✅ 完成（耗時 {time.time() - t0:.0f} 秒）：{out_path}")
    subprocess.run(['open', str(out_path)])

if __name__ == '__main__':
    main()
