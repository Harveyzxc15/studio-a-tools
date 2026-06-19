#!/usr/bin/env python3
# TOOL: 門市日報 本日銷售/休假人數（北一＋北二）
"""
掃 Apple Mail 收件匣裡各門市寄來的「門市Daily report」，
從附件 0XX_<店>日報_日期.xlsx 抓「本日銷售人數」「本日休假人數」。

比對方式：用附件檔名開頭的「門市代碼」(004/009/...) 對應門市，
不依賴主旨文字（主旨空格、Daily/DailyReport 等寫法不一都不影響）。

可獨立執行（python3 daily_headcount.py），也可被 arpedia_daily.py 引用
（呼叫 print_headcount()）。
"""
import glob, os, email, email.header, io, re, unicodedata
from datetime import datetime, timedelta
import openpyxl

MAIL_BASE = os.path.expanduser('~/Library/Mail/V10')

# 門市代碼 = 附件檔名開頭三碼。北一(6) + 北二(7)
CODES = {
    '北一區': {'士林': '004', '微風': '005', '美麗華': '024',
              '阿波羅': '046', '高島屋': '054', '羅東': '057'},
    '北二區': {'永和': '009', '板橋誠品': '025', '西門': '050', '花蓮': '055',
              '板橋遠百': '063', '新莊宏匯': '064', '新店裕隆城': '068'},
}
OUR_CODES = {c for d in CODES.values() for c in d.values()}


def w(s):
    """字串顯示寬度（全形字算 2）。"""
    return sum(2 if unicodedata.east_asian_width(ch) in 'WF' else 1 for ch in str(s))


def pad(s, n, align='<'):
    s = str(s)
    sp = ' ' * max(0, n - w(s))
    return s + sp if align == '<' else sp + s


def dec(s):
    if not s:
        return ''
    return ''.join(t.decode(c or 'utf-8', 'ignore') if isinstance(t, bytes) else t
                   for t, c in email.header.decode_header(s))


def _bar(done, total, width=24):
    filled = int(width * done / total) if total else width
    return '[' + '█' * filled + '·' * (width - filled) + f'] {done}/{total}'


_DEMO_RE = re.compile(r'[Dd]emo\s*次數\s*[：:]\s*(\d+)')   # 容各種寫法，避開「Demo 照片」


def _body_text(msg):
    """取信件純文字內文。"""
    parts = msg.walk() if msg.is_multipart() else [msg]
    for p in parts:
        if p.get_content_type() == 'text/plain':
            try:
                return p.get_payload(decode=True).decode(
                    p.get_content_charset() or 'utf-8', 'ignore')
            except Exception:
                continue
    return ''


def _parse_demo(msg):
    """從內文抓 ARpedia Demo 次數，找不到回 None。"""
    m = _DEMO_RE.search(_body_text(msg))
    return int(m.group(1)) if m else None


def collect_reports(since_days=4, progress=True):
    """掃收件匣，回傳 {(date_token, code): {'xlsx': bytes, 'demo': int|None}}。
    只收北一/北二 13 家。"""
    cut = (datetime.now() - timedelta(days=since_days)).timestamp()
    files = [f for f in glob.glob(
                os.path.join(MAIL_BASE, '*/INBOX.mbox/**/Messages/*.emlx'),
                recursive=True)
             if os.path.getmtime(f) >= cut]
    total = len(files)
    found = {}
    for i, f in enumerate(files, 1):
        if progress and (i % 10 == 0 or i == total):
            print(f"\r  掃描信件 {_bar(i, total)}", end='', flush=True)
        try:
            with open(f, 'rb') as fh:
                raw = fh.read()
            msg = email.message_from_bytes(raw.split(b'\n', 1)[1])
            subj = dec(msg.get('Subject', ''))
            if 'aily' not in subj:                      # Daily / daily / DailyReport
                continue
            for p in msg.walk():
                fn = dec(p.get_filename() or '')
                if not fn.endswith('.xlsx') or '日報' not in fn or '維修中心' in fn:
                    continue
                mc = re.match(r'(\d{3})_', fn)          # 檔名開頭門市代碼
                md = re.search(r'(20\d{6})', fn)        # 檔名裡的日期
                if not mc or not md or mc.group(1) not in OUR_CODES:
                    continue
                key = (md.group(1), mc.group(1))
                if key not in found:                    # 同一封重複下載只取一次
                    found[key] = {'xlsx': p.get_payload(decode=True),
                                  'demo': _parse_demo(msg)}
        except Exception:
            continue
    if progress and total:
        print(f"\r  掃描信件 {_bar(total, total)}  完成")
    return found


def read_headcount(data):
    """從 xlsx bytes 讀 (本日銷售人數, 本日休假人數)，找標籤抓右邊一格。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    sale = off = None
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            if cell.value == '本日銷售人數':
                sale = ws.cell(cell.row, cell.column + 1).value
            elif cell.value == '本日休假人數':
                off = ws.cell(cell.row, cell.column + 1).value
    return sale, off


def _i(v):
    return int(v) if isinstance(v, (int, float)) else 0


def print_headcount(region=None):
    """印出門市本日銷售/休假人數明細 + 小計。
    region=None 印北一＋北二（含總計）；給 '北一區'/'北二區' 只印該區。"""
    regions = [region] if region in CODES else list(CODES)
    label = region if region in CODES else '北一＋北二'
    print(f"\n【門市日報 本日銷售/休假人數、ARpedia Demo（{label}）】")
    reports = collect_reports()
    if not reports:
        print("  ⚠️  收件匣近 4 天找不到門市 Daily report")
        return
    target = max(t for t, _ in reports)                 # 最新一份日報日期
    d = datetime.strptime(target, '%Y%m%d')
    print(f"  資料日期：{d.strftime('%Y-%m-%d')}\n")

    def hdr():
        return (f"    {pad('門市', 12)}{pad('本日銷售人數', 14, '>')}"
                f"{pad('本日休假人數', 14, '>')}{pad('ARpedia Demo', 14, '>')}")

    g_sale = g_off = g_demo = 0
    for reg in regions:
        if len(regions) > 1:
            print(f"  ＜{reg}＞")
        print(hdr())
        r_sale = r_off = r_demo = 0
        for store, code in CODES[reg].items():
            rec = reports.get((target, code))
            if rec is None:
                print(f"    {pad(store, 12)}{pad('未收到', 14, '>')}"
                      f"{pad('', 14, '>')}{pad('', 14, '>')}")
                continue
            sale, off = read_headcount(rec['xlsx'])
            sale, off = _i(sale), _i(off)
            demo = rec['demo']
            r_sale += sale
            r_off += off
            r_demo += _i(demo)
            demo_s = '未填' if demo is None else demo
            print(f"    {pad(store, 12)}{pad(sale, 14, '>')}"
                  f"{pad(off, 14, '>')}{pad(demo_s, 14, '>')}")
        print(f"    {pad('小計', 12)}{pad(r_sale, 14, '>')}"
              f"{pad(r_off, 14, '>')}{pad(r_demo, 14, '>')}\n")
        g_sale += r_sale
        g_off += r_off
        g_demo += r_demo
    if len(regions) > 1:
        print(f"  {pad('總計（北一＋北二）', 14)}{pad(g_sale, 12, '>')}"
              f"{pad(g_off, 14, '>')}{pad(g_demo, 14, '>')}")


if __name__ == '__main__':
    print_headcount()
